"""深交所ETF全量导入 - 从Excel导入所有深交所ETF份额"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from datetime import date
from app.core.database import SessionLocal, init_db
from app.models.models import ETFShare

init_db()

files = [
    os.path.expanduser("~/LocalFile/ChromeDownloads/深交所ETF_20240930_20250303.xlsx"),
    os.path.expanduser("~/LocalFile/ChromeDownloads/深交所ETF_20250304_20250402.xlsx"),
    os.path.expanduser("~/LocalFile/ChromeDownloads/深交所ETF_20250403_20250429.xlsx"),
    os.path.expanduser("~/LocalFile/ChromeDownloads/深交所ETF_20250430_20250930.xlsx"),
    os.path.expanduser("~/LocalFile/ChromeDownloads/深交所ETF_20251021_20260403.xlsx"),
]

# 读取全部Excel
all_data = {}  # {code: {date_str: shares_yi}}
for f in files:
    wb = openpyxl.load_workbook(f)
    ws = wb.active
    for r in range(2, ws.max_row + 1):
        code = str(ws.cell(r, 2).value or "").strip()
        dt = str(ws.cell(r, 1).value).strip()[:10]
        shares_raw = float(ws.cell(r, 4).value or 0)
        all_data.setdefault(code, {})[dt] = round(shares_raw / 1e8, 4)
    print(f"  {f.split('/')[-1]}: 已读取")

print(f"\n深交所ETF总数: {len(all_data)} 只")
all_dates = set()
for d in all_data.values():
    all_dates.update(d.keys())
all_dates = sorted(all_dates)
print(f"日期范围: {all_dates[0]} ~ {all_dates[-1]}, 共 {len(all_dates)} 个交易日")

# 写入数据库
db = SessionLocal()
all_codes = list(all_data.keys())
deleted = db.query(ETFShare).filter(ETFShare.fund_code.in_(all_codes)).delete(synchronize_session=False)
db.commit()
print(f"删除旧数据: {deleted} 条")

total = 0
for code, date_shares in all_data.items():
    for dt_str, shares in sorted(date_shares.items()):
        db.add(ETFShare(
            fund_code=code, trade_date=date.fromisoformat(dt_str),
            price=None, total_market_cap=None,
            shares=shares, change_shares=None, source="szse_excel",
        ))
        total += 1
    if total % 10000 < len(date_shares):
        db.commit()
        print(f"  进度: {total} 条...")

db.commit()
print(f"\n✅ 深交所全量导入完成: {len(all_data)} 只ETF, {total} 条记录")
db.close()
